import cv2
import pickle
import os
import csv
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

# === Fichiers du modèle, labels et Excel ===
MODEL_FILE = "modele_lbph.yml"
LABELS_FILE = "labels.pkl"
CSV_MEMBRES = "membres.csv"
EXCEL_FILE = "presence.xlsx"

# === Charger le modèle LBPH ===
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(MODEL_FILE)
print(f"[INFO] Modèle chargé depuis {MODEL_FILE}")

# === Charger les labels ===
with open(LABELS_FILE, "rb") as f:
    labels = pickle.load(f)
# labels = {"matricule": id} ; on inverse pour id -> matricule
id_to_matricule = {v: k for k, v in labels.items()}

# === Charger les membres depuis membres.csv ===
membres_info = {}
if not os.path.exists(CSV_MEMBRES):
    print(f"[ERREUR] Fichier {CSV_MEMBRES} introuvable !")
    exit()
with open(CSV_MEMBRES, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        matricule = row["Matricule"].strip()
        nom = row["Nom Complet"].strip()
        departement = row["Département"].strip()
        membres_info[matricule] = {"nom": nom, "departement": departement}

# === Initialisation caméra et détecteur ===
cap = cv2.VideoCapture(0)
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
cam_name = "Caméra principale"

# === Création ou chargement du fichier Excel ===
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Présences"
    ws.append([
        "Matricule", "Nom complet", "Département", "Date",
        "Heure arrivée", "Heure départ", "Score confiance",
        "Statut", "Caméra", "Commentaire"
    ])
    wb.save(EXCEL_FILE)
else:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

# === Dictionnaire pour éviter répétitions rapides ===
derniere_reco = {}

# === Fonction pour trouver la ligne du jour ===
def presence_du_jour(matricule):
    today = datetime.now().date()
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(matricule):
            date_val = row[3].value
            if isinstance(date_val, datetime):
                date_val = date_val.date()
            if date_val == today:
                return row
    return None

# === Fonction pour déterminer le statut ===
def calcul_statut(arrivee_time, seuil="08:00:00"):
    seuil_time = datetime.strptime(seuil, "%H:%M:%S").time()
    if arrivee_time.time() <= seuil_time:
        return "✅ Présent"
    else:
        return "⏳ Retard"

# === Fonction pour enregistrer présence ===
def enregistrer_presence(matricule, nom, departement, score):
    now = datetime.now()
    if matricule in derniere_reco and now - derniere_reco[matricule] < timedelta(seconds=30):
        return

    ligne = presence_du_jour(matricule)
    statut = calcul_statut(now)

    if ligne is None:
        # Nouvelle entrée
        ws.append([
            matricule, nom, departement, now.date(),
            now.strftime("%H:%M:%S"), "", score,
            statut, cam_name, ""
        ])
        print(f"{nom} ({matricule}) arrivé à {now.strftime('%H:%M:%S')} - {statut}")
    else:
        # Départ
        depart_val = ligne[5].value
        if depart_val in (None, ""):
            ligne[5].value = now.strftime("%H:%M:%S")
            ligne[6].value = score
            # Mettre à jour le statut si nécessaire
            arrivee_val = ligne[4].value
            if isinstance(arrivee_val, str):
                arrivee_time = datetime.strptime(arrivee_val, "%H:%M:%S")
            else:
                arrivee_time = arrivee_val
            ligne[7].value = calcul_statut(arrivee_time)
            print(f"{nom} ({matricule}) parti à {now.strftime('%H:%M:%S')}")

    wb.save(EXCEL_FILE)
    derniere_reco[matricule] = now

# === Boucle principale ===
print("Reconnaissance faciale en temps réel. Appuyez sur 'q' pour quitter.")

while True:
    ret, frame = cap.read()
    if not ret:
        print("[ERREUR] Impossible d'accéder à la caméra.")
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.2, 5)

    for (x, y, w, h) in faces:
        visage = gray[y:y+h, x:x+w]
        visage_redim = cv2.resize(visage, (100, 100))

        label_id, confiance = recognizer.predict(visage_redim)
        matricule = id_to_matricule.get(label_id, "00000")
        membre = membres_info.get(matricule, {"nom": "Inconnu", "departement": "Inconnu"})
        nom_complet = membre["nom"]
        departement = membre["departement"]
        score_confiance = round(confiance, 2)

        if confiance < 80:
            enregistrer_presence(matricule, nom_complet, departement, score_confiance)
            couleur = (0, 255, 0)
        else:
            nom_complet = "Inconnu"
            couleur = (0, 0, 255)

        # Affichage
        cv2.rectangle(frame, (x, y), (x+w, y+h), couleur, 2)
        cv2.putText(frame, f"{nom_complet} ({matricule}) {int(confiance)}", (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, couleur, 2)

    cv2.imshow("Reconnaissance Faciale", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
print(f"[INFO] Présences sauvegardées dans {EXCEL_FILE}")
